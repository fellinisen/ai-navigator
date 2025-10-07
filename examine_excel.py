import pandas as pd
import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('Assessment.xlsx')
print('Available sheets:', wb.sheetnames)

# Examine the Maturity vs. Target Gaps sheet
if 'Maturity vs. Target Gaps' in wb.sheetnames:
    gaps_sheet = wb['Maturity vs. Target Gaps']
    print('\nFirst 20 rows of Maturity vs. Target Gaps sheet:')
    for i, row in enumerate(gaps_sheet.iter_rows(values_only=True), 1):
        # Filter out None values for cleaner output
        filtered_row = [cell for cell in row if cell is not None]
        if filtered_row:  # Only print non-empty rows
            print(f'Row {i}: {filtered_row}')
        if i >= 20:
            break
    
    # Try to find the actual header row
    print('\n\nLooking for header patterns:')
    for i, row in enumerate(gaps_sheet.iter_rows(values_only=True), 1):
        row_values = [str(cell).strip() if cell is not None else "" for cell in row]
        if any("tier" in val.lower() for val in row_values):
            print(f'Row {i} (potential header): {row_values}')
        if i >= 10:
            break
            
    # Try reading with different header rows
    print('\n\nTrying to read with header=1:')
    try:
        df = pd.read_excel('Assessment.xlsx', sheet_name='Maturity vs. Target Gaps', header=1)
        print(df.head(10))
        print('\nColumn names:', df.columns.tolist())
    except Exception as e:
        print(f'Error reading with header=1: {e}')
        
else:
    print('Maturity vs. Target Gaps sheet not found')