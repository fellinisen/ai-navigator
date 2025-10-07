from fastapi import FastAPI, HTTPException, BackgroundTasks, Request
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, EmailStr
from typing import Dict, List, Any, Optional
import json
import os
import pandas as pd
import re
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from pathlib import Path
from openpyxl import load_workbook
from motor.motor_asyncio import AsyncIOMotorClient

app = FastAPI(title="AI Navigator Assessment", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class QuestionnaireResponse(BaseModel):
    questionnaire_id: str
    responses: Dict[str, Any]
    target_responses: Dict[str, Any] = {}
    user_email: EmailStr


class AssessmentResult(BaseModel):
    questionnaire_id: str
    score: float
    category: str
    recommendations: List[str]
    detailed_results: Dict[str, Any]
    maturity_results: Optional[Dict[str, Any]] = None


class QuestionOption(BaseModel):
    value: float
    text: str
    score: float


class Question(BaseModel):
    id: str
    text: str
    type: str
    options: List[QuestionOption]
    weight: float = 1.0
    category: Optional[str] = None
    tier1: Optional[str] = None
    tier2: Optional[str] = None
    theme: Optional[str] = None


class Questionnaire(BaseModel):
    id: str
    title: str
    description: str
    questions: List[Question]
    thresholds: Optional[Dict[str, Any]] = None


DATA_DIR = Path(__file__).parent / "data"
QUESTIONNAIRES_JSON = DATA_DIR / "questionnaires.json"

questionnaire_data: Dict[str, Any] = {}
mongo_client: Optional[AsyncIOMotorClient] = None
mongo_db = None
responses_collection = None


def load_questionnaires_from_json():
    global questionnaire_data
    if QUESTIONNAIRES_JSON.exists():
        try:
            with QUESTIONNAIRES_JSON.open("r", encoding="utf-8") as f:
                questionnaire_data = json.load(f)
        except Exception:
            questionnaire_data = {}





def load_excel_data_once():
    global questionnaire_data
    try:
        excel_path = os.path.join(os.path.dirname(__file__), "..", "Assessment.xlsx")
        excel_path = os.path.abspath(excel_path)
        if not os.path.exists(excel_path):
            load_sample_data()
            return

        excel_file = pd.ExcelFile(excel_path)
        # Use openpyxl for robust header detection on the Assessment sheet
        wb = load_workbook(excel_path, data_only=True)
        assessment_ws = wb["Assessment"] if "Assessment" in wb.sheetnames else None
        gaps_sheet = (
            pd.read_excel(excel_file, sheet_name="Maturity vs. Target Gaps")
            if "Maturity vs. Target Gaps" in excel_file.sheet_names
            else None
        )

        questionnaire_data = {
            "ai_readiness": {
                "id": "ai_readiness",
                "title": "AI Readiness Assessment",
                "description": "Evaluate your organization's readiness for AI implementation",
                "questions": [],
                "targets": {"tier2": {}, "tier1": {}},
            }
        }

        if assessment_ws is not None:

            def _norm(s: str) -> str:
                return re.sub(r"[^a-z0-9]+", "", str(s).lower())

            header_row_idx = None
            # Find the header row within the first 25 rows
            for r in range(1, min(25, assessment_ws.max_row) + 1):
                values = [
                    str(c.value).strip() if c.value is not None else ""
                    for c in assessment_ws[r]
                ]
                norms = [_norm(v) for v in values]
                if any(
                    "tier3" in n
                    and "capability" in n
                    and ("description" in n or "score" in n)
                    for n in norms
                ):
                    header_row_idx = r
                    break

            if header_row_idx:
                headers = [
                    str(c.value).strip() if c.value is not None else ""
                    for c in assessment_ws[header_row_idx]
                ]
                norms = [_norm(h) for h in headers]

                def find_idx(*preds):
                    for i, n in enumerate(norms, start=1):
                        if all(p in n for p in preds):
                            return i
                    return None

                idx_t1 = find_idx("tier1")
                idx_t2 = find_idx("tier2")
                # Column 4 is "Tier 3 (the capability to score)" - this is the question text
                idx_q_desc = find_idx("tier3", "capability", "score") or find_idx(
                    "tier3thecapabilitytoscore"
                )
                # Column 5 is "Tier 3 (capability description)" - this is the description/theme
                idx_q_theme = find_idx("tier3", "capability", "description") or find_idx(
                    "tier3capabilitydescription"
                )

                fixed_options = [
                    {"value": 1, "text": "Does not exist", "score": 1.0},
                    {"value": 2, "text": "Partially exists", "score": 2.0},
                    {"value": 3, "text": "Fully exists", "score": 3.0},
                    {"value": 4, "text": "Fully exists and optimized", "score": 4.0},
                    {"value": 5, "text": "Fully exists and adaptive", "score": 5.0},
                ]

                # Track current tier values as we iterate through rows
                current_tier1 = ""
                current_tier2 = ""
                seen_tier2_categories = set()  # Track which Tier 2 categories we've already added
                
                for r in range(header_row_idx + 1, assessment_ws.max_row + 1):
                    # Get values from all relevant columns
                    q_text = (
                        assessment_ws.cell(row=r, column=idx_q_desc).value
                        if idx_q_desc
                        else None
                    )
                    theme = (
                        assessment_ws.cell(row=r, column=idx_q_theme).value
                        if idx_q_theme
                        else None
                    )
                    t1_val = (
                        assessment_ws.cell(row=r, column=idx_t1).value
                        if idx_t1
                        else None
                    )
                    t2_val = (
                        assessment_ws.cell(row=r, column=idx_t2).value
                        if idx_t2
                        else None
                    )
                    
                    # Update current tier values when they change (cascade down)
                    # Check both Tier 1 columns (short name and descriptive)
                    if t1_val and str(t1_val).strip() and str(t1_val).strip().lower() != "none":
                        current_tier1 = str(t1_val).strip()
                    else:
                        # Also check column 2 (descriptive Tier 1)
                        t1_desc_val = (
                            assessment_ws.cell(row=r, column=2).value
                            if idx_t1
                            else None
                        )
                        if t1_desc_val and str(t1_desc_val).strip() and str(t1_desc_val).strip().lower() != "none":
                            current_tier1 = str(t1_desc_val).strip()
                    
                    if t2_val and str(t2_val).strip() and str(t2_val).strip().lower() != "none":
                        current_tier2 = str(t2_val).strip()
                    
                    # Only create questions for rows with actual question text (Tier 3)
                    if q_text is None or str(q_text).strip() == "":
                        continue
                    
                    # Skip if we've already added a question for this Tier 2 category
                    if current_tier2 in seen_tier2_categories:
                        continue
                    
                    # Mark this Tier 2 category as seen
                    seen_tier2_categories.add(current_tier2)

                    question = {
                        "id": f"q_{r}",
                        "text": str(q_text).strip(),
                        "type": "multiple_choice",
                        "options": fixed_options,
                        "weight": 1.0,
                        "category": str(theme).strip() if theme else None,
                        "tier1": current_tier1,
                        "tier2": current_tier2,
                        "tier3": str(q_text).strip(),  # Tier 3 is the individual capability/question
                        "theme": str(theme).strip() if theme else None,
                    }
                    questionnaire_data["ai_readiness"]["questions"].append(question)

        # Parse targets (best-effort) from "Maturity vs. Target Gaps"
        if gaps_sheet is not None:
            cols = {str(c).strip(): c for c in gaps_sheet.columns}
            col_t2 = next((c for k, c in cols.items() if k.lower() == "tier 2"), None)
            col_t1 = next((c for k, c in cols.items() if k.lower() == "tier 1"), None)
            # Try multiple target column names
            col_target = next(
                (c for k, c in cols.items() if "target" in k.lower()),
                None,
            )
            if col_t2 and col_target:
                for _, row in gaps_sheet.iterrows():
                    t2 = row.get(col_t2)
                    target = row.get(col_target)
                    if pd.notna(t2) and pd.notna(target):
                        questionnaire_data["ai_readiness"]["targets"]["tier2"][
                            str(t2).strip()
                        ] = float(target)
            if col_t1 and col_target:
                for _, row in gaps_sheet.iterrows():
                    t1 = row.get(col_t1)
                    target = row.get(col_target)
                    if pd.notna(t1) and pd.notna(target):
                        questionnaire_data["ai_readiness"]["targets"]["tier1"][
                            str(t1).strip()
                        ] = float(target)

        # Do NOT fallback to sample data; keep as-is so issues can be diagnosed

        # Persist parsed questionnaire for future runs
        save_questionnaires_to_json()
        print(f"Successfully loaded {len(questionnaire_data.get('ai_readiness', {}).get('questions', []))} questions")
    except Exception as e:
        # Print the exception for debugging
        print(f"Exception in load_excel_data_once: {e}")
        import traceback
        traceback.print_exc()
        # Keep existing in-memory data; avoid overwriting with sample
        pass


load_questionnaires_from_json()
# Check if we have actual questionnaire content, not just an empty dict
if not questionnaire_data or not questionnaire_data.get("ai_readiness", {}).get("questions"):
    load_excel_data_once()
# Init MongoDB if configured
mongodb_url = os.environ.get("MONGODB_URL")
mongodb_db = os.environ.get("MONGODB_DB", "ai_navigator")
responses_collection = None

if mongodb_url:
    try:
        mongo_client = AsyncIOMotorClient(mongodb_url)
        mongo_db = mongo_client[mongodb_db]
        responses_collection = mongo_db["responses"]
        print(f"Connected to MongoDB: {mongodb_url}")
    except Exception as e:
        print(f"Failed to connect to MongoDB: {e}")
else:
    print("No MongoDB URL provided, skipping database connection")


def _ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def save_questionnaires_to_json():
    _ensure_data_dir()
    with QUESTIONNAIRES_JSON.open("w", encoding="utf-8") as f:
        json.dump(questionnaire_data, f, ensure_ascii=False, indent=2)


def _norm(s):
    return re.sub(r"[^a-z0-9]+", "", str(s).lower())


def save_questionnaires_to_json():
    _ensure_data_dir()
    with QUESTIONNAIRES_JSON.open("w", encoding="utf-8") as f:
        json.dump(questionnaire_data, f, ensure_ascii=False, indent=2)


def load_sample_data():
    global questionnaire_data
    questionnaire_data = {
        "ai_readiness": {
            "id": "ai_readiness",
            "title": "AI Readiness Assessment",
            "description": "Evaluate your organization's readiness for AI implementation",
            "questions": [
                {
                    "id": "q_1",
                    "text": "How would you rate your organization's current data management capabilities?",
                    "type": "scale",
                    "options": [
                        {"value": 1, "text": "Poor", "score": 1},
                        {"value": 2, "text": "Fair", "score": 2},
                        {"value": 3, "text": "Good", "score": 3},
                        {"value": 4, "text": "Very Good", "score": 4},
                        {"value": 5, "text": "Excellent", "score": 5},
                    ],
                    "weight": 1.5,
                },
                {
                    "id": "q_2",
                    "text": "Does your organization have a dedicated AI strategy?",
                    "type": "multiple_choice",
                    "options": [
                        {"value": 0, "text": "No strategy exists", "score": 0},
                        {"value": 1, "text": "Strategy is being developed", "score": 2},
                        {
                            "value": 2,
                            "text": "Strategy exists but not implemented",
                            "score": 3,
                        },
                        {
                            "value": 3,
                            "text": "Strategy is partially implemented",
                            "score": 4,
                        },
                        {
                            "value": 4,
                            "text": "Strategy is fully implemented",
                            "score": 5,
                        },
                    ],
                    "weight": 2.0,
                },
                {
                    "id": "q_3",
                    "text": "What is your organization's experience with AI technologies?",
                    "type": "multiple_choice",
                    "options": [
                        {"value": 0, "text": "No experience", "score": 0},
                        {"value": 1, "text": "Limited pilot projects", "score": 2},
                        {
                            "value": 2,
                            "text": "Several successful implementations",
                            "score": 4,
                        },
                        {"value": 3, "text": "Extensive AI deployment", "score": 5},
                    ],
                    "weight": 1.8,
                },
            ],
        }
    }


def calculate_assessment_score(
    questionnaire_id: str,
    responses: Dict[str, Any],
    target_responses: Dict[str, Any] = {},
) -> AssessmentResult:
    questionnaire = questionnaire_data.get(questionnaire_id)
    if not questionnaire:
        raise HTTPException(status_code=404, detail="Questionnaire not found")

    total_score = 0.0
    max_possible_score = 0.0
    detailed_results: Dict[str, Any] = {}
    # Aggregation by tiers - collect Tier 3 (individual question) scores
    tier3_scores: Dict[str, List[float]] = {}  # tier3 -> [scores]
    tier2_to_tier3: Dict[str, List[str]] = {}  # tier2 -> [tier3 names]
    tier1_to_tier2: Dict[str, List[str]] = {}  # tier1 -> [tier2 names]
    
    # For target maturity
    tier3_target_scores: Dict[str, List[float]] = {}
    tier2_target_scores: Dict[str, List[float]] = {}
    tier1_target_scores: Dict[str, List[float]] = {}

    for question in questionnaire["questions"]:
        qid = question["id"]
        if qid in responses:
            response_value = responses[qid]
            question_score = 0.0
            for option in question["options"]:
                if option["value"] == response_value:
                    question_score = float(option["score"])
                    break
            weight = float(question.get("weight", 1.0))
            weighted_score = question_score * weight
            total_score += weighted_score

            max_question_score = (
                max([float(opt["score"]) for opt in question["options"]]) * weight
            )
            max_possible_score += max_question_score

            # Get target maturity if provided
            target_score = None
            if target_responses and qid in target_responses:
                target_value = target_responses[qid]
                for option in question["options"]:
                    if option["value"] == target_value:
                        target_score = float(option["score"])
                        break

            detailed_results[qid] = {
                "question": question["text"],
                "response": response_value,
                "score": question_score,
                "weighted_score": weighted_score,
                "tier1": question.get("tier1"),
                "tier2": question.get("tier2"),
                "tier3": question.get("tier3"),
                "theme": question.get("theme"),
                "target_score": target_score,
            }

            # Collect for tier aggregations (use raw scores 1-5)
            t1 = question.get("tier1")
            t2 = question.get("tier2")
            t3 = question.get("tier3")
            
            # Store Tier 3 scores (individual questions)
            if t3:
                tier3_scores.setdefault(t3, []).append(question_score)
                if target_score:
                    tier3_target_scores.setdefault(t3, []).append(target_score)
            
            # Build tier hierarchy mappings
            if t1 and t2:
                if t2 not in tier1_to_tier2.get(t1, []):
                    tier1_to_tier2.setdefault(t1, []).append(t2)
            if t2 and t3:
                if t3 not in tier2_to_tier3.get(t2, []):
                    tier2_to_tier3.setdefault(t2, []).append(t3)

    percentage_score = (
        (total_score / max_possible_score * 100.0) if max_possible_score > 0 else 0.0
    )

    if percentage_score >= 80:
        category = "AI Ready"
        recommendations = [
            "Strong AI readiness",
            "Consider advanced AI implementations",
            "Focus on scaling existing initiatives",
        ]
    elif percentage_score >= 60:
        category = "Moderately Ready"
        recommendations = [
            "Good foundation for AI",
            "Address identified gaps",
            "Run pilot programs",
        ]
    elif percentage_score >= 40:
        category = "Developing Readiness"
        recommendations = [
            "Improve data management and strategy",
            "Invest in AI education",
            "Start with small projects",
        ]
    else:
        category = "Not Ready"
        recommendations = [
            "Build basic data infrastructure",
            "Develop AI awareness and strategy",
            "Focus on foundational capabilities",
        ]

    # Build maturity results similar to spreadsheet outputs
    # First calculate Tier 2 averages from Tier 3 scores
    tier2_results = []
    tier2_scores = {}  # For Tier 1 calculation
    
    for tier2_name, tier3_list in tier2_to_tier3.items():
        # Collect all Tier 3 scores for this Tier 2
        all_tier3_scores = []
        all_tier3_targets = []
        
        for tier3_name in tier3_list:
            if tier3_name in tier3_scores:
                all_tier3_scores.extend(tier3_scores[tier3_name])
            if tier3_name in tier3_target_scores:
                all_tier3_targets.extend(tier3_target_scores[tier3_name])
        
        current = sum(all_tier3_scores) / len(all_tier3_scores) if all_tier3_scores else 0.0
        target = sum(all_tier3_targets) / len(all_tier3_targets) if all_tier3_targets else None
        
        # Use spreadsheet targets as fallback
        if target is None:
            target = questionnaire.get("targets", {}).get("tier2", {}).get(tier2_name)
        
        gap = (target - current) if (target is not None) else None
        
        tier2_results.append({
            "name": tier2_name,
            "current_maturity": current,
            "target_maturity": target,
            "gap": gap,
        })
        
        # Store for Tier 1 calculation
        tier2_scores[tier2_name] = current
        if target is not None:
            tier2_target_scores[tier2_name] = [target]

    # Then calculate Tier 1 averages from Tier 2 scores
    tier1_results = []
    
    for tier1_name, tier2_list in tier1_to_tier2.items():
        # Collect all Tier 2 scores for this Tier 1
        all_tier2_scores = []
        all_tier2_targets = []
        
        for tier2_name in tier2_list:
            if tier2_name in tier2_scores:
                all_tier2_scores.append(tier2_scores[tier2_name])
            if tier2_name in tier2_target_scores and tier2_target_scores[tier2_name]:
                all_tier2_targets.extend(tier2_target_scores[tier2_name])
        
        current = sum(all_tier2_scores) / len(all_tier2_scores) if all_tier2_scores else 0.0
        target = sum(all_tier2_targets) / len(all_tier2_targets) if all_tier2_targets else None
        
        # Use spreadsheet targets as fallback
        if target is None:
            target = questionnaire.get("targets", {}).get("tier1", {}).get(tier1_name)
        
        gap = (target - current) if (target is not None) else None
        
        tier1_results.append({
            "name": tier1_name,
            "current_maturity": current,
            "target_maturity": target,
            "gap": gap,
        })

    result = AssessmentResult(
        questionnaire_id=questionnaire_id,
        score=percentage_score,
        category=category,
        recommendations=recommendations,
        detailed_results=detailed_results,
    )
    # Attach extended maturity details
    result_dict = result.dict()

    # Create maturity vs target plot data for tier1 sections
    maturity_plot_data = []
    for item in tier1_results:
        if (
            item["name"]
            and item["current_maturity"] is not None
            and item["target_maturity"] is not None
        ):
            maturity_plot_data.append(
                {
                    "name": item["name"],
                    "current_maturity": item["current_maturity"],
                    "target_maturity": item["target_maturity"],
                }
            )

    result_dict["maturity_results"] = {
        "tier1": tier1_results,
        "tier2": tier2_results,
        "maturity_plot": maturity_plot_data,
        "overall_maturity": (
            (
                sum([s for scores in tier3_scores.values() for s in scores])
                / max(1, sum([len(scores) for scores in tier3_scores.values()]))
            )
            if tier3_scores
            else 0.0
        ),
    }
    return AssessmentResult(**result_dict)


async def send_email(email: str, assessment_result: AssessmentResult):
    try:
        msg = MIMEMultipart()
        msg["From"] = settings.SMTP_USER
        msg["To"] = email
        msg["Subject"] = (
            f"Your {assessment_result.questionnaire_id.replace('_', ' ').title()} Results"
        )

        results_html = "".join(
            [
                f"<p><strong>{res['question']}</strong><br>Score: {res['score']} (weighted: {res['weighted_score']:.2f})</p>"
                for _, res in assessment_result.detailed_results.items()
            ]
        )

        html_body = f"""
        <html>
        <body style="font-family: Arial, sans-serif; max-width: 600px; margin: 0 auto;">
            <div style="background-color: #171C8F; color: white; padding: 20px; text-align: center;">
                <h1>Assessment Results</h1>
            </div>
            <div style="padding: 20px;">
                <h2 style="color: #171C8F;">Your Score: {assessment_result.score:.1f}%</h2>
                <h3 style="color: #0072CE;">Category: {assessment_result.category}</h3>
                <h3 style="color: #171C8F;">Recommendations:</h3>
                <ul>
                    {''.join([f'<li>{rec}</li>' for rec in assessment_result.recommendations])}
                </ul>
                <h3 style="color: #171C8F;">Detailed Results:</h3>
                <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px;">
                    {results_html}
                </div>
            </div>
            <div style="background-color: #f0f0f0; padding: 15px; text-align: center; margin-top: 20px;">
                <p style="color: #666;">Thank you for using AI Navigator Assessment Tool</p>
            </div>
        </body>
        </html>
        """

        msg.attach(MIMEText(html_body, "html"))
        with smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT) as server:
            server.starttls()
            server.login(settings.SMTP_USER, settings.SMTP_PASSWORD)
            server.send_message(msg)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to send email: {e}")


@app.get("/")
async def root():
    return {"message": "AI Navigator Assessment API"}


@app.get("/questionnaires")
async def get_questionnaires():
    return {
        "questionnaires": [
            {
                "id": qid,
                "title": data["title"],
                "description": data["description"],
            }
            for qid, data in questionnaire_data.items()
        ]
    }


@app.get("/questionnaires/{questionnaire_id}")
async def get_questionnaire(questionnaire_id: str):
    if questionnaire_id not in questionnaire_data:
        raise HTTPException(status_code=404, detail="Questionnaire not found")
    return questionnaire_data[questionnaire_id]


@app.get("/debug/excel-columns")
async def debug_excel_columns():
    """Expose detected headers and a small sample to help troubleshoot parsing."""
    try:
        excel_path = os.path.join(os.path.dirname(__file__), "..", "Assessment.xlsx")
        excel_path = os.path.abspath(excel_path)
        if not os.path.exists(excel_path):
            raise HTTPException(status_code=404, detail="Assessment.xlsx not found")

        wb = load_workbook(excel_path, data_only=True)
        ws_name = "Assessment"
        if ws_name not in wb.sheetnames:
            raise HTTPException(status_code=404, detail=f"Sheet '{ws_name}' not found")
        ws = wb[ws_name]

        def _norm(s: str) -> str:
            return re.sub(r"[^a-z0-9]+", "", str(s).lower())

        # Capture the first 30 rows as text for inspection
        preview_rows = []
        for r in range(1, min(30, ws.max_row) + 1):
            row_vals = [
                str(c.value).strip() if c.value is not None else "" for c in ws[r]
            ]
            preview_rows.append(row_vals)

        # Find header row as in parser
        header_row_idx = None
        for r in range(1, min(25, ws.max_row) + 1):
            values = [
                str(c.value).strip() if c.value is not None else "" for c in ws[r]
            ]
            norms = [_norm(v) for v in values]
            if any(
                "tier3" in n
                and "capability" in n
                and ("description" in n or "score" in n)
                for n in norms
            ):
                header_row_idx = r
                break

        headers = (
            [
                str(c.value).strip() if c.value is not None else ""
                for c in ws[header_row_idx]
            ]
            if header_row_idx
            else []
        )
        norms = [_norm(h) for h in headers] if headers else []
        return {
            "excel_path": excel_path,
            "header_row_index": header_row_idx,
            "headers": headers,
            "normalized_headers": norms,
            "preview_first_rows": preview_rows,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Debug failed: {e}")


def _require_admin(request: Request):
    admin_key = request.headers.get("X-Admin-Key")
    if not admin_key or admin_key != settings.SECRET_KEY:
        raise HTTPException(status_code=403, detail="Forbidden: invalid admin key")


@app.post("/admin/questionnaires")
async def create_questionnaire(q: Questionnaire, request: Request):
    _require_admin(request)
    if q.id in questionnaire_data:
        raise HTTPException(status_code=400, detail="Questionnaire ID already exists")
    questionnaire_data[q.id] = q.dict()
    save_questionnaires_to_json()
    return {"message": "Questionnaire created", "id": q.id}


@app.put("/admin/questionnaires/{questionnaire_id}")
async def update_questionnaire(
    questionnaire_id: str, q: Questionnaire, request: Request
):
    _require_admin(request)
    if questionnaire_id not in questionnaire_data:
        raise HTTPException(status_code=404, detail="Questionnaire not found")
    if q.id != questionnaire_id:
        # Allow renaming by removing old and inserting new
        questionnaire_data.pop(questionnaire_id)
    questionnaire_data[q.id] = q.dict()
    save_questionnaires_to_json()
    return {"message": "Questionnaire updated", "id": q.id}


@app.delete("/admin/questionnaires/{questionnaire_id}")
async def delete_questionnaire(questionnaire_id: str, request: Request):
    _require_admin(request)
    if questionnaire_id not in questionnaire_data:
        raise HTTPException(status_code=404, detail="Questionnaire not found")
    questionnaire_data.pop(questionnaire_id)
    save_questionnaires_to_json()
    return {"message": "Questionnaire deleted", "id": questionnaire_id}


@app.post("/submit-assessment")
async def submit_assessment(
    response: QuestionnaireResponse, background_tasks: BackgroundTasks
):
    assessment_result = calculate_assessment_score(
        response.questionnaire_id, response.responses, response.target_responses
    )
    background_tasks.add_task(send_email, response.user_email, assessment_result)
    # Persist to Mongo if available
    if responses_collection is not None:
        try:
            doc = {
                "created_at": datetime.utcnow().isoformat(),
                "questionnaire_id": response.questionnaire_id,
                "user_email": str(response.user_email),
                "responses": response.responses,
                "target_responses": response.target_responses,
                "results": assessment_result.dict(),
            }
            await responses_collection.insert_one(doc)
        except Exception as e:
            print(f"Failed to save response: {e}")
    else:
        print("No responses collection available, skipping persistence")
    return {
        "message": "Assessment submitted successfully",
        "results": assessment_result.dict(),
        "email_sent": True,
    }


@app.get("/admin/responses")
async def list_responses(request: Request, limit: int = 50, skip: int = 0):
    _require_admin(request)
    if responses_collection is None:
        raise HTTPException(status_code=500, detail="Responses storage not initialized")
    cursor = responses_collection.find({}, skip=skip, limit=limit).sort(
        "created_at", -1
    )
    items = []
    async for doc in cursor:
        doc_id = str(doc.get("_id")) if doc.get("_id") else None
        res = doc.get("results", {}) if isinstance(doc.get("results"), dict) else {}
        maturity = res.get("maturity_results", {}) if isinstance(res, dict) else {}
        items.append(
            {
                "id": doc_id,
                "created_at": doc.get("created_at"),
                "questionnaire_id": doc.get("questionnaire_id"),
                "user_email": doc.get("user_email"),
                "score": res.get("score"),
                "tier1": maturity.get("tier1"),
                "tier2": maturity.get("tier2"),
            }
        )
    return {"responses": items}


@app.get("/admin/responses/{response_id}")
async def get_response(response_id: str, request: Request):
    _require_admin(request)
    if responses_collection is None:
        raise HTTPException(status_code=500, detail="Responses storage not initialized")
    try:
        from bson import ObjectId

        doc = await responses_collection.find_one({"_id": ObjectId(response_id)})
        if not doc:
            raise HTTPException(status_code=404, detail="Response not found")
        doc["_id"] = str(doc["_id"])  # stringify
        return doc
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid response id")


@app.get("/health")
async def health_check():
    """Health check endpoint for deployment platforms"""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}


if os.path.exists(os.path.join(os.path.dirname(__file__), "..", "frontend", "dist")):
    app.mount(
        "/",
        StaticFiles(
            directory=os.path.join(os.path.dirname(__file__), "..", "frontend", "dist"),
            html=True,
        ),
        name="static",
    )
