import os
import pandas as pd
from openpyxl import load_workbook
import json

def test_tier1_fix():
    """Test with the same logic as backend main.py"""
    
    print("=== Testing Tier 1 Fix with Backend Logic ===\n")
    
    # Initialize questionnaire data
    questionnaire_data = {
        "ai_readiness": {
            "id": "ai_readiness",
            "title": "AI Readiness Assessment",
            "description": "Evaluate your organization's readiness for AI implementation",
            "questions": [],
            "targets": {"tier2": {}, "tier1": {}},
        }
    }
    
    try:
        excel_path = "Assessment.xlsx"
        wb = load_workbook(excel_path, data_only=True)
        assessment_ws = wb["Assessment"]
        
        # Find header row (same logic as backend)
        def _norm(s):
            return "".join(c.lower() for c in str(s) if c.isalnum())
        
        header_row_idx = None
        for r in range(1, 11):
            values = [
                str(c.value).strip() if c.value is not None else ""
                for c in assessment_ws[r]
            ]
            norms = [_norm(v) for v in values]
            if any(
                "tier3" in n
                and "capability" in n
                and ("description" in n or "score" in n)
                for n in norms
            ):
                header_row_idx = r
                break
        
        if not header_row_idx:
            print("❌ Header row not found")
            return
        
        print(f"✅ Header row found at line {header_row_idx}")
        
        # Get column indices (same logic as backend)
        headers = [
            str(c.value).strip() if c.value is not None else ""
            for c in assessment_ws[header_row_idx]
        ]
        norms = [_norm(h) for h in headers]

        def find_idx(*preds):
            for i, n in enumerate(norms, start=1):
                if all(p in n for p in preds):
                    return i
            return None

        idx_t1 = find_idx("tier1")
        idx_t2 = find_idx("tier2")
        idx_q_desc = find_idx("tier3", "capability", "score") or find_idx(
            "tier3thecapabilitytoscore"
        )
        idx_q_theme = find_idx("tier3", "capability", "description") or find_idx(
            "tier3capabilitydescription"
        )
        
        print(f"Column indices: T1={idx_t1}, T2={idx_t2}, Q_desc={idx_q_desc}, Q_theme={idx_q_theme}")
        
        if None in [idx_t1, idx_t2, idx_q_desc, idx_q_theme]:
            print("❌ Missing required columns")
            return
        
        # Fixed options
        fixed_options = [
            {"value": 1, "text": "Does not exist", "score": 1.0},
            {"value": 2, "text": "Partially exists", "score": 2.0},
            {"value": 3, "text": "Fully exists", "score": 3.0},
            {"value": 4, "text": "Fully exists and optimized", "score": 4.0},
            {"value": 5, "text": "Fully exists and adaptive", "score": 5.0},
        ]
        
        # Process rows (same logic as backend)
        seen_tier2_categories = set()
        current_tier1 = current_tier2 = None
        questions_added = 0
        
        for r in range(header_row_idx + 1, assessment_ws.max_row + 1):
            q_text = (
                assessment_ws.cell(row=r, column=idx_q_desc).value
                if idx_q_desc
                else None
            )
            q_theme = (
                assessment_ws.cell(row=r, column=idx_q_theme).value
                if idx_q_theme
                else None
            )
            t1_val = (
                assessment_ws.cell(row=r, column=idx_t1).value
                if idx_t1
                else None
            )
            t2_val = (
                assessment_ws.cell(row=r, column=idx_t2).value
                if idx_t2
                else None
            )
            
            # Update current tier values when they change (cascade down)
            # Check both Tier 1 columns (short name and descriptive)
            if t1_val and str(t1_val).strip() and str(t1_val).strip().lower() != "none":
                current_tier1 = str(t1_val).strip()
            else:
                # Also check column 2 (descriptive Tier 1)
                t1_desc_val = (
                    assessment_ws.cell(row=r, column=2).value
                    if idx_t1
                    else None
                )
                if t1_desc_val and str(t1_desc_val).strip() and str(t1_desc_val).strip().lower() != "none":
                    current_tier1 = str(t1_desc_val).strip()
            
            if t2_val and str(t2_val).strip() and str(t2_val).strip().lower() != "none":
                current_tier2 = str(t2_val).strip()
            
            # Only create questions for rows with actual question text (Tier 3)
            if q_text is None or str(q_text).strip() == "":
                continue
            
            # Skip if we've already added a question for this Tier 2 category
            if current_tier2 in seen_tier2_categories:
                continue
            
            # Mark this Tier 2 category as seen
            seen_tier2_categories.add(current_tier2)

            question = {
                "id": f"q_{r}",
                "text": str(q_text).strip(),
                "type": "scale",
                "options": fixed_options,
                "tier1": current_tier1,
                "tier2": current_tier2,
                "tier3": str(q_text).strip(),
                "theme": str(q_theme).strip() if q_theme else "",
            }
            
            questionnaire_data["ai_readiness"]["questions"].append(question)
            questions_added += 1
            
            if questions_added <= 10:
                print(f"Question {questions_added}: T1='{current_tier1}', T2='{current_tier2}'")
                print(f"  Text: {str(q_text).strip()[:60]}...")
        
        print(f"\n✅ Successfully loaded {questions_added} questions")
        
        # Analyze Tier distribution
        tier1_counts = {}
        tier2_counts = {}
        for q in questionnaire_data["ai_readiness"]["questions"]:
            tier1 = q.get("tier1", "Unknown")
            tier2 = q.get("tier2", "Unknown")
            tier1_counts[tier1] = tier1_counts.get(tier1, 0) + 1
            tier2_counts[tier2] = tier2_counts.get(tier2, 0) + 1
        
        print(f"Tier 1 categories: {len(tier1_counts)}")
        for t1, count in tier1_counts.items():
            print(f"  - {t1}: {count} questions")
        
        print(f"Tier 2 categories: {len(tier2_counts)}")
        
        # Save to JSON
        with open("questionnaires_tier1_fixed.json", "w", encoding="utf-8") as f:
            json.dump(questionnaire_data, f, ensure_ascii=False, indent=2)
        print(f"✅ Saved questionnaire data to questionnaires_tier1_fixed.json")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_tier1_fix()